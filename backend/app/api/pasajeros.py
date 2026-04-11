from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user_required
from app.models.usuario import Usuario
from app.schemas.pasajero import PasajeroCreate, PasajeroUpdate, PasajeroResponse
from app.services.permisos_service import get_user_proyectos, can_access_proyecto
from app.services.pasajero_service import (
    get_pasajeros,
    get_pasajero_by_id,
    create_pasajero,
    update_pasajero,
    delete_pasajero,
)
from app.services.proyecto_service import get_proyecto_by_id
from app.services.ruta_service import get_ruta_by_id
from app.services.tipo_pasajero_service import get_tipo_pasajero_by_id

router = APIRouter()


def _validate_tipo_pasajero_proyecto(db: Session, tipo_pasajero_id: str, proyecto_id: str) -> None:
    tp = get_tipo_pasajero_by_id(db, tipo_pasajero_id)
    if not tp or tp.proyecto_id != proyecto_id:
        raise HTTPException(status_code=404, detail="Tipo de pasajero no encontrado")


def _to_response(db: Session, obj) -> PasajeroResponse:
    d = PasajeroResponse.model_validate(obj).model_dump()
    d["tiene_contrasena"] = bool(getattr(obj, "password_hash", None))
    if getattr(obj, "creado_por", None):
        u = db.query(Usuario).filter(Usuario.usuario_id == obj.creado_por).first()
        d["creado_por_nombre"] = u.nombre_usuario if u else obj.creado_por
    if getattr(obj, "actualizado_por", None):
        u = db.query(Usuario).filter(Usuario.usuario_id == obj.actualizado_por).first()
        d["actualizado_por_nombre"] = u.nombre_usuario if u else obj.actualizado_por
    return PasajeroResponse(**d)


@router.get("", response_model=List[PasajeroResponse])
def list_pasajeros(
    proyecto_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_required),
):
    allowed = get_user_proyectos(db, current_user.usuario_id)
    return [_to_response(db, p) for p in get_pasajeros(db, proyecto_id=proyecto_id, allowed_proyecto_ids=allowed, skip=skip, limit=limit)]


@router.post("", response_model=PasajeroResponse, status_code=status.HTTP_201_CREATED)
def crear_pasajero(
    pasajero: PasajeroCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_required),
):
    if not get_proyecto_by_id(db, pasajero.proyecto_id):
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not can_access_proyecto(db, current_user.usuario_id, pasajero.proyecto_id):
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if pasajero.ruta_id and not get_ruta_by_id(db, pasajero.ruta_id):
        raise HTTPException(status_code=404, detail="Ruta no encontrada")
    if pasajero.tipo_pasajero_id:
        _validate_tipo_pasajero_proyecto(db, pasajero.tipo_pasajero_id, pasajero.proyecto_id)
    p = create_pasajero(db, pasajero, creado_por_id=current_user.usuario_id)
    return _to_response(db, p)


@router.get("/{pasajero_id}", response_model=PasajeroResponse)
def obtener_pasajero(
    pasajero_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_required),
):
    p = get_pasajero_by_id(db, pasajero_id)
    if not p:
        raise HTTPException(status_code=404, detail="Pasajero no encontrado")
    if not can_access_proyecto(db, current_user.usuario_id, p.proyecto_id):
        raise HTTPException(status_code=404, detail="Pasajero no encontrado")
    return _to_response(db, p)


@router.put("/{pasajero_id}", response_model=PasajeroResponse)
def actualizar_pasajero(
    pasajero_id: str,
    pasajero: PasajeroUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_required),
):
    p = get_pasajero_by_id(db, pasajero_id)
    if not p or not can_access_proyecto(db, current_user.usuario_id, p.proyecto_id):
        raise HTTPException(status_code=404, detail="Pasajero no encontrado")
    upd = pasajero.model_dump(exclude_unset=True)
    if upd.get("tipo_pasajero_id"):
        _validate_tipo_pasajero_proyecto(db, upd["tipo_pasajero_id"], p.proyecto_id)
    updated = update_pasajero(db, pasajero_id, pasajero, actualizado_por_id=current_user.usuario_id)
    if not updated:
        raise HTTPException(status_code=404, detail="Pasajero no encontrado")
    return _to_response(db, updated)


@router.delete("/{pasajero_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_pasajero(
    pasajero_id: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_required),
):
    p = get_pasajero_by_id(db, pasajero_id)
    if not p or not can_access_proyecto(db, current_user.usuario_id, p.proyecto_id):
        raise HTTPException(status_code=404, detail="Pasajero no encontrado")
    if not delete_pasajero(db, pasajero_id):
        raise HTTPException(status_code=404, detail="Pasajero no encontrado")


@router.post("/importar")
def importar_pasajeros(
    proyecto_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_required),
):
    """Importa pasajeros desde Excel o CSV. Columnas: cedula, nombre, direccion (opc), lat/latitud, lng/longitud (opc), contrasena (opc), ruta_id (opc), horario_habitual (opc), placa_asignada (opc)."""
    if not get_proyecto_by_id(db, proyecto_id):
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not can_access_proyecto(db, current_user.usuario_id, proyecto_id):
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    try:
        import io
        import csv
        content = file.file.read()
        if file.filename and file.filename.lower().endswith(".csv"):
            decoded = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2):
                    rows.append(dict(zip(headers, [str(c.value or "").strip() for c in row])))
            except ImportError:
                raise HTTPException(status_code=501, detail="Instalar openpyxl para importar Excel: pip install openpyxl")
        from decimal import Decimal, InvalidOperation

        def _cell_decimal(row_dict, *keys):
            for k in keys:
                raw = row_dict.get(k)
                if raw is None:
                    continue
                s = str(raw).strip().replace(",", ".")
                if not s:
                    continue
                try:
                    return Decimal(s)
                except InvalidOperation:
                    continue
            return None

        creados = 0
        for row in rows:
            cedula = row.get("cedula", row.get("cédula", "")).strip()
            nombre = row.get("nombre", "").strip()
            if not cedula or not nombre:
                continue
            from app.schemas.pasajero import PasajeroCreate

            lat = _cell_decimal(row, "lat", "latitud")
            lng = _cell_decimal(row, "lng", "longitud", "lon")
            pw = row.get("contrasena", row.get("password", "")).strip() or None
            create_pasajero(db, PasajeroCreate(
                proyecto_id=proyecto_id,
                cedula=cedula,
                nombre=nombre,
                direccion=row.get("direccion", row.get("dirección", "")).strip() or None,
                lat=lat,
                lng=lng,
                contrasena=pw,
                ruta_id=row.get("ruta_id", "").strip() or None,
                horario_habitual=row.get("horario_habitual", row.get("horario", "")).strip() or None,
                placa_asignada=row.get("placa_asignada", row.get("placa", "")).strip() or None,
            ), creado_por_id=current_user.usuario_id)
            creados += 1
        return {"message": f"Importados {creados} pasajeros"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
